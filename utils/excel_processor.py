import os
import shutil
import json
import yaml
from datetime import datetime
from typing import List, Dict, Any, Optional
from utils.log_handler import LogHandler
from openpyxl import load_workbook
import xlrd  # 添加 xlrd 导入

class ExcelProcessor:
    """Excel处理器，负责处理不同供应商的送货单"""
    
    def __init__(self):
        """初始化Excel处理器"""
        self.logger = LogHandler().get_logger('ExcelProcessor')
        self.config = self._load_config()
        
    def _load_config(self) -> dict:
        """加载配置文件"""
        config_path = os.path.join("config", "processor_config.yaml")
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
                self.logger.info("成功加载配置文件")
                return config
        except Exception as e:
            self.logger.error("加载配置文件失败: %s", LogHandler.format_error(e))
            return {}
            
    def _save_json(self, data: List[Dict[str, Any]], filename: str, supplier: str):
        """
        保存JSON数据到指定位置
        
        参数:
            data: 要保存的数据
            filename: JSON文件名
            supplier: 供应商标识
        """
        try:
            # 确保输出目录存在
            output_dir = self.config['paths'][supplier]['json_output']
            os.makedirs(output_dir, exist_ok=True)
            
            # 构建完整的文件路径
            json_path = os.path.join(output_dir, filename)
            
            # 保存JSON文件
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
                
            self.logger.info("JSON数据已保存到: %s", json_path)
            return json_path
        except Exception as e:
            self.logger.error("保存JSON数据失败: %s", LogHandler.format_error(e))
            return None
            
    def _move_excel(self, excel_path: str, supplier: str):
        """
        移动Excel文件到指定位置
        
        参数:
            excel_path: Excel文件路径
            supplier: 供应商标识
        """
        try:
            # 确保归档目录存在
            archive_dir = self.config['paths'][supplier]['excel_archive']
            os.makedirs(archive_dir, exist_ok=True)
            
            # 构建目标路径
            filename = os.path.basename(excel_path)
            target_path = os.path.join(archive_dir, filename)
            
            # 移动文件
            shutil.move(excel_path, target_path)
            self.logger.info("Excel文件已移动到: %s", target_path)
            return True
        except Exception as e:
            self.logger.error("移动Excel文件失败: %s", LogHandler.format_error(e))
            return False
            
    def _format_date(self, date_str: str, from_format: bool = True) -> Optional[str]:
        """
        日期格式转换
        
        Args:
            date_str: 日期字符串
            from_format: True表示转换为YYYY-MM-DD格式，False表示转换为YYYYMMDD格式
            
        Returns:
            Optional[str]: 转换后的日期字符串，如果转换失败则返回None
        """
        try:
            # 处理特殊情况
            if date_str in ["0000-00-00", "00000000"]:
                return "0000-00-00" if from_format else "00000000"
                
            if from_format:
                # 将其他格式转换为YYYY-MM-DD
                date_formats = [
                    '%Y%m%d',     # YYYYMMDD
                    '%Y-%m-%d',   # YYYY-MM-DD
                    '%Y/%m/%d',   # YYYY/MM/DD
                    '%Y.%m.%d',   # YYYY.MM.DD
                    '%Y年%m月%d日' # YYYY年MM月DD日
                ]
                
                for fmt in date_formats:
                    try:
                        date_obj = datetime.strptime(date_str, fmt)
                        return date_obj.strftime('%Y-%m-%d')
                    except ValueError:
                        continue
            else:
                # 将YYYY-MM-DD转换为YYYYMMDD（用于比较和存储）
                try:
                    if '-' in date_str:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                        return date_obj.strftime('%Y%m%d')
                    else:
                        date_obj = datetime.strptime(date_str, '%Y%m%d')
                        return date_str
                except ValueError:
                    pass
                    
            self.logger.warning(f"无法解析日期格式: {date_str}")
            return None
            
        except Exception as e:
            self.logger.error(f"处理日期时出错: {str(e)}")
            return None
            
    def _compare_dates(self, date1: str, date2: str) -> int:
        """
        比较两个日期的大小
        
        Args:
            date1: 第一个日期（YYYY-MM-DD格式）
            date2: 第二个日期（YYYY-MM-DD格式）
            
        Returns:
            int: 如果date1 > date2返回1，如果date1 < date2返回-1，如果相等返回0
        """
        try:
            # 处理特殊情况
            if date1 == date2:
                return 0
                
            if date1 == "0000-00-00":
                return -1
                
            if date2 == "0000-00-00":
                return 1
                
            # 转换为YYYYMMDD格式进行比较
            date1_fmt = self._format_date(date1, False)
            date2_fmt = self._format_date(date2, False)
            
            if not date1_fmt or not date2_fmt:
                return 0
                
            if date1_fmt > date2_fmt:
                return 1
            elif date1_fmt < date2_fmt:
                return -1
            else:
                return 0
                
        except Exception as e:
            self.logger.error(f"比较日期时出错: {str(e)}")
            return 0
            
    def _validate_and_format_data(self, data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        验证和格式化数据，确保符合配置文件中定义的格式
        
        Args:
            data: 原始数据字典
            
        Returns:
            Optional[Dict[str, Any]]: 格式化后的数据字典，如果验证失败则返回None
        """
        try:
            # 获取字段定义
            fields = self.config['json_format']['fields']
            formatted_data = {}
            
            # 验证每个字段
            for field in fields:
                field_name = field['name']
                field_type = field['type']
                required = field['required']
                
                # 获取字段值
                value = data.get(field_name)
                
                # 检查必填字段
                if required and value is None:
                    self.logger.error(f"缺少必填字段: {field_name}")
                    return None
                    
                # 如果字段不存在且非必填，使用默认值
                if value is None:
                    formatted_data[field_name] = ""
                    continue
                    
                # 根据字段类型进行格式化
                try:
                    if field_type == "date":
                        # 确保日期格式为YYYY-MM-DD
                        formatted_data[field_name] = self._format_date(str(value))
                    elif field_type == "integer":
                        # 确保数字字段为整数
                        formatted_data[field_name] = int(float(value)) if value else 0
                    elif field_type == "string":
                        # 确保字符串字段为字符串类型，并去除首尾空白
                        formatted_data[field_name] = str(value).strip() if value else ""
                    else:
                        formatted_data[field_name] = value
                except Exception as e:
                    self.logger.error(f"字段 {field_name} 格式化失败: {str(e)}")
                    return None
                    
            return formatted_data
            
        except Exception as e:
            self.logger.error(f"数据验证和格式化失败: {str(e)}")
            return None
            
    def process_excel(self, download_path: str, rule_name: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        处理指定目录下的所有Excel文件并返回按日期组织的数据字典。
        
        Args:
            download_path: Excel文件所在目录
            rule_name: 规则名称，用于确定使用哪个供应商的处理逻辑
            
        Returns:
            Dict[str, List[Dict[str, Any]]]: 返回按日期组织的数据字典
                - 键: 日期字符串(YYYY-MM-DD格式)
                - 值: 该日期对应的数据记录列表
        """
        try:
            self.logger.info(f"开始处理目录: {download_path}")
            
            # 获取供应商标识
            supplier = rule_name.split("_")[0]
            
            # 用于存储所有数据
            all_data: Dict[str, List[Dict[str, Any]]] = {}
            
            # 确保目录存在
            if not os.path.exists(download_path):
                self.logger.error(f"目录不存在: {download_path}")
                return {}
                
            # 遍历目录下的所有文件
            for filename in os.listdir(download_path):
                file_path = os.path.join(download_path, filename)
                
                # 跳过目录
                if os.path.isdir(file_path):
                    continue
                    
                # 检查文件扩展名
                if not filename.lower().endswith(('.xls', '.xlsx')):
                    continue
                    
                try:
                    self.logger.info(f"处理文件: {filename}")
                    
                    # 根据供应商选择相应的处理方法
                    if "池州华宇" in rule_name:
                        data_dict = self._process_huayu_return_dict(file_path)
                    elif "山东汉旗" in rule_name:
                        data_dict = self._process_hanqi_return_dict(file_path)
                    elif "江苏芯丰" in rule_name:
                        data_dict = self._process_xinfeng_return_dict(file_path)
                    else:
                        self.logger.error("未知的供应商类型")
                        continue
                        
                    # 验证和格式化数据
                    for date, data_list in data_dict.items():
                        formatted_list = []
                        for item in data_list:
                            formatted_item = self._validate_and_format_data(item)
                            if formatted_item:
                                formatted_list.append(formatted_item)
                                
                        if formatted_list:
                            if date in all_data:
                                all_data[date].extend(formatted_list)
                            else:
                                all_data[date] = formatted_list
                            
                    # 处理完成后，将文件移动到归档目录
                    self._move_excel(file_path, supplier)
                    
                except Exception as e:
                    self.logger.error(f"处理文件 {filename} 失败: {str(e)}")
                    continue
                    
            # 保存每个日期的数据到对应的JSON文件
            if all_data:
                json_output_dir = self.config['paths'][supplier]['json_output']
                os.makedirs(json_output_dir, exist_ok=True)
                
                for date, data_list in all_data.items():
                    try:
                        json_filename = f"{supplier}送货单_{date}.json"
                        json_path = os.path.join(json_output_dir, json_filename)
                        
                        with open(json_path, 'w', encoding='utf-8') as f:
                            json.dump(data_list, f, ensure_ascii=False, indent=2)
                            
                        self.logger.info(f"已保存JSON数据到: {json_path}")
                        
                    except Exception as e:
                        self.logger.error(f"保存JSON数据失败: {str(e)}")
                        continue
                        
            return all_data
            
        except Exception as e:
            self.logger.error(f"处理Excel文件失败: {str(e)}")
            return {}
            
    def _process_huayu_return_dict(self, excel_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        处理池州华宇的送货单Excel文件并返回数据字典
        
        处理说明:
        1. 固定读取'Page 1'工作表
        2. 日期位置固定在P4单元格
        3. 从第7行开始读取数据，直到遇到'TOTAL'行
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            Dict[str, List[Dict[str, Any]]]: 按日期组织的数据字典
        """
        try:
            # 加载Excel工作簿，data_only=True表示读取值而不是公式
            wb = load_workbook(excel_path, data_only=True)
            sheet = wb['Page 1']  # 固定使用Page 1工作表
            
            data_dict = {}
            # 从固定位置(P4)获取日期
            date_str = sheet['P4'].value
            delivery_date = self._format_date(str(date_str))
            
            if not delivery_date:
                self.logger.error("无法获取送货日期，跳过处理")
                return {}
                
            data_list = []
            
            # 从第8行开始遍历数据，直到遇到TOTAL行
            for row in range(8, sheet.max_row + 1):
                # 检查是否到达表格末尾（TOTAL行）
                if sheet[f'N{row}'].value == 'TOTAL':
                    break
                    
                # 跳过空行（以订单号是否存在为判断依据）
                if not sheet[f'D{row}'].value:
                    continue
                    
                try:
                    # 提取每行数据并构建数据字典
                    row_data = {
                        "送货日期": delivery_date,
                        "订单号": sheet[f'D{row}'].value,
                        "品名": sheet[f'F{row}'].value,
                        "封装形式": sheet[f'K{row}'].value,
                        "打印批号": sheet[f'J{row}'].value,
                        "数量": int(sheet[f'Q{row}'].value or 0),  # 如果为空则默认为0
                        "晶圆名称": sheet[f'O{row}'].value,
                        "晶圆批号": sheet[f'L{row}'].value,
                        "供应商": "池州华宇"
                    }
                    data_list.append(row_data)
                except Exception as e:
                    self.logger.error(f"处理第 {row} 行数据时出错: {str(e)}")
                    continue
                    
            # 如果有数据，则添加到返回字典中
            if data_list:
                data_dict[delivery_date] = data_list
                
            return data_dict
            
        except Exception as e:
            self.logger.error(f"处理池州华宇送货单失败: {str(e)}")
            return {}
            
    def _get_last_process_date(self, supplier: str) -> str:
        """
        获取供应商最后一次处理的送货日期
        
        Args:
            supplier: 供应商标识
            
        Returns:
            str: 最后处理的日期（YYYYMMDD格式），如果没有记录则返回'00000000'
        """
        try:
            # 读取记录文件
            process_dates_file = os.path.join("config", "process_dates.json")
            if not os.path.exists(process_dates_file):
                # 如果文件不存在，创建默认内容
                default_content = {"山东汉旗": "0000-00-00", "池州华宇": "0000-00-00", "江苏芯丰": "0000-00-00"}
                with open(process_dates_file, 'w', encoding='utf-8') as f:
                    json.dump(default_content, f, ensure_ascii=False, indent=2)
                return "0000-00-00"
                
            # 读取日期记录
            with open(process_dates_file, 'r', encoding='utf-8') as f:
                dates = json.load(f)
                return dates.get(supplier, "0000-00-00")
                
        except Exception as e:
            self.logger.error(f"获取最后处理日期失败: {str(e)}")
            return "0000-00-00"
            
    def _update_last_process_date(self, supplier: str, date: str) -> bool:
        """
        更新供应商最后一次处理的送货日期
        
        Args:
            supplier: 供应商标识
            date: 新的处理日期（YYYY-MM-DD格式）
            
        Returns:
            bool: 更新成功返回True，失败返回False
        """
        try:
            process_dates_file = os.path.join("config", "process_dates.json")
            
            # 读取现有记录
            with open(process_dates_file, 'r', encoding='utf-8') as f:
                dates = json.load(f)
                
            # 更新日期
            current_date = dates.get(supplier, "0000-00-00")
            # 转换为YYYYMMDD格式进行比较
            if self._format_date(date, False) > self._format_date(current_date, False):
                dates[supplier] = date
                
                # 保存更新后的记录
                with open(process_dates_file, 'w', encoding='utf-8') as f:
                    json.dump(dates, f, ensure_ascii=False, indent=2)
                    
                self.logger.info(f"已更新{supplier}的最后处理日期: {date}")
                return True
                
            return False
            
        except Exception as e:
            self.logger.error(f"更新最后处理日期失败: {str(e)}")
            return False
            
    def _is_xls_file(self, file_path: str) -> bool:
        """
        检查文件是否为旧版 .xls 格式
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 如果是 .xls 文件返回 True，否则返回 False
        """
        return file_path.lower().endswith('.xls')

    def _process_hanqi_return_dict(self, excel_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        处理山东汉旗的送货单Excel文件并返回数据字典
        
        处理说明:
        1. 遍历所有工作表
        2. 每个工作表的日期位置在G3单元格，格式为"日期:YYYY-MM-DD"
        3. 从第6行开始读取数据，直到遇到'Total'行
        4. 只处理日期大于上次处理日期的数据
        5. 支持旧版 .xls 和新版 .xlsx 格式
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            Dict[str, List[Dict[str, Any]]]: 按日期组织的数据字典
        """
        try:
            # 获取最后处理日期
            last_process_date = self._get_last_process_date("山东汉旗")
            self.logger.info(f"山东汉旗最后处理日期: {last_process_date}")
            
            data_dict = {}
            max_processed_date = last_process_date  # 用于记录本次处理的最大日期

            if self._is_xls_file(excel_path):
                # 使用 xlrd 处理 .xls 文件
                workbook = xlrd.open_workbook(excel_path)
                
                # 遍历所有工作表
                for sheet in workbook.sheets():
                    # 获取日期单元格内容 (G3 对应的是 row=2, col=6)
                    date_cell = sheet.cell_value(2, 6)
                    
                    # 检查日期单元格格式是否正确
                    if not date_cell or '日期：' not in str(date_cell):
                        continue
                        
                    # 提取并转换日期
                    delivery_date = self._format_date(str(date_cell).split('日期：')[-1].strip())
                    if not delivery_date:
                        continue
                        
                    # 检查日期是否大于最后处理日期
                    if self._compare_dates(delivery_date, last_process_date) <= 0:
                        self.logger.info(f"跳过已处理的日期: {delivery_date}")
                        continue
                        
                    # 更新最大处理日期
                    if self._compare_dates(delivery_date, max_processed_date) > 0:
                        max_processed_date = delivery_date
                        
                    data_list = []
                    
                    # 从第7行开始读取数据 (索引从0开始，所以是6)
                    for row in range(6, sheet.nrows):
                        # 检查是否到达表格末尾（Total行）
                        if 'Total' in str(sheet.cell_value(row, 7)):  # H列对应索引7
                            break
                        # 跳过空行
                        if not sheet.cell_value(row, 4):  # E列对应索引4
                            continue
                            
                        try:
                            # 提取每行数据
                            row_data = {
                                "送货日期": delivery_date,
                                "订单号": str(sheet.cell_value(row, 4)),  # E列
                                "品名": str(sheet.cell_value(row, 2)),    # C列
                                "封装形式": str(sheet.cell_value(row, 7)), # H列
                                "打印批号": str(sheet.cell_value(row, 5)), # F列
                                "数量": int(float(sheet.cell_value(row, 8)) or 0),  # I列
                                "晶圆名称": str(sheet.cell_value(row, 1)), # B列
                                "晶圆批号": str(sheet.cell_value(row, 3)), # D列
                                "供应商": "山东汉旗"
                            }
                            data_list.append(row_data)
                        except Exception as e:
                            self.logger.error(f"处理第 {row + 1} 行数据时出错: {str(e)}")
                            continue
                            
                    # 合并相同日期的数据
                    if data_list:
                        if delivery_date in data_dict:
                            data_dict[delivery_date].extend(data_list)
                        else:
                            data_dict[delivery_date] = data_list
            else:
                # 使用 openpyxl 处理 .xlsx 文件
                wb = load_workbook(excel_path, data_only=True)
                
                # 遍历所有工作表
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    # 获取日期单元格内容
                    date_cell = sheet['G3'].value
                    
                    # 检查日期单元格格式是否正确
                    if not date_cell or '日期:' not in str(date_cell):
                        continue
                        
                    # 提取并转换日期
                    delivery_date = self._format_date(str(date_cell).split('日期:')[-1].strip())
                    if not delivery_date:
                        continue
                        
                    # 检查日期是否大于最后处理日期
                    if self._compare_dates(delivery_date, last_process_date) <= 0:
                        self.logger.info(f"跳过已处理的日期: {delivery_date}")
                        continue
                        
                    # 更新最大处理日期
                    if self._compare_dates(delivery_date, max_processed_date) > 0:
                        max_processed_date = delivery_date
                        
                    data_list = []
                    
                    # 从第6行开始读取数据
                    for row in range(6, sheet.max_row + 1):
                        # 检查是否到达表格末尾（Total行）
                        if sheet[f'H{row}'].value == 'Total':
                            break
                            
                        # 跳过空行
                        if not sheet[f'E{row}'].value:
                            continue
                            
                        try:
                            # 提取每行数据
                            row_data = {
                                "送货日期": delivery_date,
                                "订单号": sheet[f'E{row}'].value,
                                "品名": sheet[f'C{row}'].value,
                                "封装形式": sheet[f'H{row}'].value,
                                "打印批号": sheet[f'F{row}'].value,
                                "数量": int(sheet[f'I{row}'].value or 0),
                                "晶圆名称": sheet[f'B{row}'].value,
                                "晶圆批号": sheet[f'D{row}'].value,
                                "供应商": "山东汉旗"
                            }
                            data_list.append(row_data)
                        except Exception as e:
                            self.logger.error(f"处理第 {row} 行数据时出错: {str(e)}")
                            continue
                            
                    # 合并相同日期的数据
                    if data_list:
                        if delivery_date in data_dict:
                            data_dict[delivery_date].extend(data_list)
                        else:
                            data_dict[delivery_date] = data_list
                            
            # 更新最后处理日期
            if self._compare_dates(max_processed_date, last_process_date) > 0:
                self._update_last_process_date("山东汉旗", max_processed_date)
                self.logger.info(f"更新山东汉旗最后处理日期为: {max_processed_date}")
                
            return data_dict
            
        except Exception as e:
            self.logger.error(f"处理山东汉旗送货单失败: {str(e)}")
            return {}
            
    def _process_xinfeng_return_dict(self, excel_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        处理江苏芯丰的送货单Excel文件并返回数据字典
        
        处理说明:
        1. 遍历所有工作表
        2. 每个工作表的日期位置在L2单元格，格式为"出货日期:YYYY-MM-DD"
        3. 从第9行开始读取数据，直到遇到空行
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            Dict[str, List[Dict[str, Any]]]: 按日期组织的数据字典
        """
        try:
            # 加载Excel工作簿
            wb = load_workbook(excel_path, data_only=True)
            data_dict = {}
            
            # 遍历所有工作表
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # 获取日期单元格内容
                date_cell = sheet['L2'].value
                
                # 检查日期单元格格式是否正确
                if not date_cell or '出货日期:' not in str(date_cell):
                    continue
                    
                # 提取并转换日期
                delivery_date = self._format_date(str(date_cell).split('出货日期:')[-1].strip())
                if not delivery_date:
                    continue
                    
                data_list = []
                
                # 从第9行开始读取数据
                for row in range(9, sheet.max_row + 1):
                    # 检查是否到达表格末尾（空行）
                    if not sheet[f'A{row}'].value:
                        break
                        
                    # 跳过空行
                    if not sheet[f'C{row}'].value:
                        continue
                        
                    try:
                        # 提取每行数据
                        row_data = {
                            "送货日期": delivery_date,
                            "订单号": sheet[f'C{row}'].value,
                            "品名": sheet[f'D{row}'].value,
                            "封装形式": sheet[f'E{row}'].value,
                            "打印批号": "",  # 芯丰没有打印批号字段
                            "数量": int(sheet[f'H{row}'].value or 0),
                            "晶圆名称": sheet[f'F{row}'].value,
                            "晶圆批号": sheet[f'G{row}'].value,
                            "供应商": "江苏芯丰"
                        }
                        data_list.append(row_data)
                    except Exception as e:
                        self.logger.error(f"处理第 {row} 行数据时出错: {str(e)}")
                        continue
                        
                # 合并相同日期的数据
                if data_list:
                    if delivery_date in data_dict:
                        data_dict[delivery_date].extend(data_list)
                    else:
                        data_dict[delivery_date] = data_list
                        
            return data_dict
            
        except Exception as e:
            self.logger.error(f"处理江苏芯丰送货单失败: {str(e)}")
            return {}